import pandas as pd
import numpy as np
import os
import re
from openpyxl import load_workbook
from utils.data_cleaning import clean_running_total
from utils.excel_utils import generate_output_dataframe
from flask import current_app


def process_excel_file(file_path):
    work_book = load_workbook(file_path)
    sheet_names = work_book.sheetnames

    # Load GVL sheet
    sheets = {}
    for s in sheet_names:
        if "GVL" in s:
            sheets["GVL"] = work_book[s]
        else:
            sheets[s] = work_book[s]

    # Read data into DataFrame
    data = [row for row in sheets["GVL"].iter_rows(values_only=True)]
    df = pd.DataFrame(data[1:], columns=data[0])

    # Clean and process data
    df["Running Total"] = clean_running_total(df["Running Total"])

    # Generate the output DataFrame
    op_df = generate_output_dataframe(df,sheets)
    print("Did the output dataframe generated")
    # Save the result to a new Excel file
    output_file = "demand_pivot_analysis.xlsx"
    output_path = os.path.join(current_app.config['OUTPUT_FOLDER'], output_file)
    op_df.to_excel(output_path, sheet_name="GVL Extracted", index=False)
    return output_file